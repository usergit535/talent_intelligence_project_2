"""
scraper/data_exporter.py

Step 4 – Validate, deduplicate, and export to:
  • output/companies.xlsx  (styled Excel with openpyxl)
  • output/companies.json  (structured JSON)
"""

from __future__ import annotations

import json
import os
import re
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config.config import EXCEL_FILE, JSON_FILE, OUTPUT_DIR
from scraper.logger import get_logger

log = get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Column ordering and human-readable labels
# ─────────────────────────────────────────────────────────────────────────────

# Internal key → export column label
FIELD_MAP = {
    "company_name":  "Company Name",
    "industry":      "Industry",
    "total_funding": "Total Funding",
    "funding_stage": "Funding Stage",
    "headquarters":  "Headquarters",
    "website":       "Website",
    "careers_page":  "Careers Page",
    "hiring_status": "Hiring Status",
    "open_jobs":     "Open Jobs",
    "job_roles":     "Hiring Roles",
    "linkedin_url":  "LinkedIn URL",
    "twitter_url":   "X/Twitter URL",
    "facebook_url":  "Facebook URL",
    "instagram_url": "Instagram URL",
    "youtube_url":   "YouTube URL",
    "description":   "Description",
    "funding_date":  "Funding Date",
    "source_url":    "Source URL",
}
COLUMNS = list(FIELD_MAP.keys())   # ordered internal keys
LABELS  = list(FIELD_MAP.values()) # matching Excel column headers

# ─────────────────────────────────────────────────────────────────────────────
# Validation helpers
# ─────────────────────────────────────────────────────────────────────────────

URL_FIELDS = {
    "website", "careers_page", "linkedin_url", "twitter_url",
    "facebook_url", "instagram_url", "youtube_url", "source_url",
}


def _valid_url(url: str) -> bool:
    if not url or url == "N/A":
        return False
    try:
        p = urlparse(url)
        return bool(p.scheme in ("http", "https") and p.netloc)
    except Exception:
        return False


def _clean_url(url) -> str:
    if not url:
        return "N/A"
    s = str(url).strip().rstrip("/")
    return s if _valid_url(s) else "N/A"


def _na(val) -> str:
    if val is None:
        return "N/A"
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "nan", "") else "N/A"


def _std_name(name: str) -> str:
    """Standardise: strip, collapse whitespace, title-case."""
    if not name or name == "N/A":
        return "N/A"
    return re.sub(r"\s+", " ", name.strip())


# ─────────────────────────────────────────────────────────────────────────────
# Main functions
# ─────────────────────────────────────────────────────────────────────────────

def validate_and_deduplicate(companies: list[dict]) -> list[dict]:
    log.info(f"[STEP 4] Validating {len(companies)} records …")
    seen_names: set[str] = set()
    cleaned: list[dict] = []

    for raw in companies:
        name = _std_name(raw.get("company_name", ""))
        key  = name.lower()
        if key in seen_names or key in ("", "n/a"):
            continue
        seen_names.add(key)

        rec: dict = {}
        for col in COLUMNS:
            val = raw.get(col, "N/A")

            if col == "company_name":
                rec[col] = name

            elif col in URL_FIELDS:
                rec[col] = _clean_url(val)

            elif col == "job_roles":
                roles = val if isinstance(val, list) else []
                rec[col] = ", ".join(roles) if roles else "N/A"

            elif col == "open_jobs":
                try:
                    rec[col] = int(val) if val else 0
                except (ValueError, TypeError):
                    rec[col] = 0

            else:
                rec[col] = _na(val)

        cleaned.append(rec)

    log.info(f"  After dedup → {len(cleaned)} companies")
    return cleaned


def export_excel(companies: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Build DataFrame with labelled columns
    rows = [{FIELD_MAP[k]: v for k, v in c.items()} for c in companies]
    df = pd.DataFrame(rows, columns=LABELS)
    df.to_excel(EXCEL_FILE, index=False, sheet_name="Companies")

    # ── Style with openpyxl ───────────────────────────────────────────────
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP        = Alignment(wrap_text=True, vertical="top")

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER

    # Data rows – alternate shading + wrap
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = WRAP
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Auto-fit column widths (capped)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(EXCEL_FILE)

    log.info(f"  Excel → {EXCEL_FILE}")
    return EXCEL_FILE


def export_json(companies: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    payload = {
        "total":     len(companies),
        "companies": [{FIELD_MAP[k]: v for k, v in c.items()} for c in companies],
    }
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    log.info(f"  JSON  → {JSON_FILE}")
    return JSON_FILE